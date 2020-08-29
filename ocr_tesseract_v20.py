# -*- coding: utf-8 -*-
"""
Created on Fri Aug 28 18:47:12 2020

@author: erik7
"""

import pandas as pd
import pytesseract
import pdf2image
import cv2
import numpy as np
import os
import re
import math
import time
#import matplotlib.pyplot as plt
import PIL
import gc
import multiprocessing
import openpyxl as op
from wand.image import Image
from skimage.metrics import structural_similarity as ssim

class imagePrep:
    
    def __init__(self):

        # assign tesseract.exe file location for Python to read from
        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
       

    def convert_from_pdf(self, file_path, img_format, img_dpi, pop_path, output_path):
        
        # convert image from PDF to 'png' file format; include path to Poppler
        return pdf2image.convert_from_path(pdf_path = file_path, fmt = img_format, dpi = img_dpi, poppler_path = pop_path, output_folder = output_path, grayscale = True, thread_count = 2)
    
       
    def remove_skew(self, img_path, save_path, page_num = False):
        
        # split the string to obtain the file name and output format to save as a new file with '_deskewed' added to the file name
        output_fmt = img_path.rsplit("\\")[-1].split(".")[1]
        output_name = img_path.rsplit("\\")[-1].split(".")[0] + "_deskewed." + output_fmt
        output_path = os.path.join(save_path, output_name)
        
        # use the 'Image' function from 'wand.image' (Image Magick library for Python) to deskew the image
        with Image(filename = img_path) as img:
            img.deskew(0.4 * img.quantum_range)       
            img.save(filename = output_path)

        return output_path
    
            
    def template_type(self, img_path):
        
        """Function to determine the kind of late scan format being read. Uses saved late scan samples of different types to determine which late scan the converted image is most similar to"""
        
        img = cv2.imread(img_path, 0)
        
        type1, type2 = cv2.imread(r"C:/Users/erik7/Documents/Late Scan Templates/template_1.png", 0), cv2.imread(r"C:/Users/erik7/Documents/Late Scan Templates/template_2.png", 0)
        
        # use 'ssim' function to test similarity between 'img' and 'type1', 'type2' templates. 'ssim' will return a score which is stored in 'results'
        results = [ssim(cv2.resize(img, (type1.shape[1], type1.shape[0])), type1, multichannel = True) # need to resize according to img to match type1 and type2 dimensions in case they differ--'ssim' function will throw error otherwise
                   , ssim(cv2.resize(img, (type2.shape[1], type2.shape[0])), type2, multichannel = True)]

        # the larger the resulting score from 'ssim' the more similar the two images. return the largest score in 'results' using the corresponding index to determine the resulting 'template_type'
        template_type = results.index(max(results)) + 1
        
        if template_type == 1:
            print("\n\tTemplate 1 detected.")
        elif template_type == 2:
            print("\n\tTemplate 2 detected.")
        else:
            print("\n\tTemplate 3 detected.")
            # template 3 formats will be any sheet that does not resemble the 'type1' or 'type2' formats. since 'type3' templates can vary in what they look like, we
            # create a catch-all for all these possible templates instead of creating their own template type instance (e.g. 'type4', 'type5', etc.)
        
        del img, type1, type2, results
        return template_type
     

    def crop_image(self, output_name, output_path, img_path, template_type):
        
        """Crops the scanned pdf image to the items table and other needed portions of the scanned image using contour detection"""
        
        def get_bounding_boxes(image):
        
            img_gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            img_blur = cv2.blur(img_gray, (15, 15))
            ret, thresh = cv2.threshold(img_blur, math.floor(np.average(img_blur)), 255, cv2.THRESH_BINARY_INV)
            dilated = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (10,10)))
    
            contours, _ = cv2.findContours(dilated, cv2.RETR_LIST,cv2.CHAIN_APPROX_SIMPLE)
            new_contours = []
            
            for contour in contours:
                if cv2.contourArea(contour) < 4000000:
                    new_contours.append(contour)
            
            del img_gray, img_blur, ret, thresh, dilated
            return [cv2.boundingRect(contour) for contour in new_contours]
        
        img = cv2.imread(img_path)
        bounding_boxes = get_bounding_boxes(img)

        # next few lines are useful for seeing all the bounding boxes/contours detected in the image
        # img_plot = img.copy()
        # for contour in new_contours:
        #     (x,y,w,h) = cv2.boundingRect(contour)
        #     #if ((w > 100) and (w < 600) and (h > 23) and (h < 80)):
        #     print(x,y,w,h)
        #     img_plot = cv2.rectangle(img_plot, (x, y), (x + w, y + h), (255, 0, 0), 2)
        #     plt.imshow(img_plot, cmap = 'gray')
        #     plt.show()

        def neighboring_cells(boundingBoxes, dimension):
            
            boxes_dict = {}
    
            # identify items table columns using position of table cells determined by x-coordinate
            if dimension == "x":
                for i in range(0, (len(boundingBoxes)-1)):
                    if boundingBoxes[i+1][0] - boundingBoxes[i][0] <= 5:
                        if bool(boxes_dict.keys()):
                            for key in boxes_dict.keys():
                                if boundingBoxes[i] in boxes_dict[key]:
                                    boxes_dict[key] = list(boxes_dict[key]) + [boundingBoxes[i+1]]
                                    break
                        else:
                            boxes_dict.setdefault(i, []).append(boundingBoxes[i])
                            boxes_dict.setdefault(i, []).append(boundingBoxes[i+1])
                    else:
                        if not (i in boxes_dict.keys()):
                            boxes_dict.setdefault(i, []).append(boundingBoxes[i])
                        boxes_dict.setdefault(i+1, []).append(boundingBoxes[i+1])
            
            # identify items table rows using position of table cells determined by y-coordinate
            elif dimension == "y":
                for i in range(0, (len(boundingBoxes)-1)):
                    if boundingBoxes[i+1][1] - boundingBoxes[i][1] <= 5:
                        if bool(boxes_dict.keys()):
                            for key in boxes_dict.keys():
                                if boundingBoxes[i] in boxes_dict[key]:
                                    boxes_dict[key] = list(boxes_dict[key]) + [boundingBoxes[i+1]]
                                    break
                        else:
                            boxes_dict.setdefault(i, []).append(boundingBoxes[i])
                            boxes_dict.setdefault(i, []).append(boundingBoxes[i+1])
                    else:
                        if not (i in boxes_dict.keys()):
                            boxes_dict.setdefault(i, []).append(boundingBoxes[i])
                        boxes_dict.setdefault(i+1, []).append(boundingBoxes[i+1])
            
            return boxes_dict


        if template_type == 1:

            # want only the bounding boxes which correspond to the cells in the items table that may actually hold information -- height is between 50 and 65 or 90 and 105 and the width is at most 675 and the x position cannot go further than 2050
            bounding_boxes = [bounding_box for bounding_box in bounding_boxes if (((bounding_box[-1] >= 50 and bounding_box[-1] <= 65) or (bounding_box[-1] >= 90 and bounding_box[-1] <= 105)) and (bounding_box[-2] <= 675) and (bounding_box[0] <= 2050))]
            # sort them -- will be sorted by x value since no key is specified
            bounding_boxes = sorted(bounding_boxes)
            # using the 'neighboring_cells' function to identify all cells that are next to each other via column detection -- this will help group all the cells found in the items table
            boxes_dict = neighboring_cells(bounding_boxes, "x")
            # filter 'boxes_dict' for keys that have at least 4 elements -- eliminates contours that may not be part of the items table by requiring that there be at least 4 contours in succession
            bounding_boxes = [boxes_dict[key] for key in boxes_dict.keys() if len(boxes_dict[key]) >= 4]
            # flatten 'bounding_boxes' to be a list of tuples instead of a list of lists with tuples
            bounding_boxes = [tup for lst in bounding_boxes for tup in lst]
            # need to repeat neighboring cells using "y" -- this is an additional filtering step to ensure only contours from the items table are used in case other contours outside the items table were found
            bounding_boxes = sorted(bounding_boxes, key = lambda tup: tup[1])
            boxes_dict = neighboring_cells(bounding_boxes, "y") 
            # filter for groups of contours that have more than two contours in succession
            bounding_boxes = [boxes_dict[key] for key in boxes_dict.keys() if len(boxes_dict[key]) > 2]
            bounding_boxes = [tup for lst in bounding_boxes for tup in lst]
            
            # next few lines are useful for seeing all the bounding boxes/contours detected after the steps above
            # img_plot = img.copy()
            # for contour in bounding_boxes:
            #     (x,y,w,h) = contour
            #     #if ((w > 100) and (w < 600) and (h > 23) and (h < 80)):
            #     print(x,y,w,h)
            #     img_plot = cv2.rectangle(img_plot, (x, y), (x + w, y + h), (255, 0, 0), 2)
            #     plt.imshow(img_plot, cmap = 'gray')
            #     plt.show()
            
            # steps above should have successfully filtered for only those cells found in the items table. 
            # get the smallest (x, y) pair and the longest (x, y) pair -- these two sets of points determine the lenght and height of the table to use for cropping
            items_x_start = sorted(bounding_boxes, key = lambda tup: tup[0])[0][0] - 15
            items_x_end = sorted(bounding_boxes, key = lambda tup: tup[0])[-1][0] + sorted(bounding_boxes, key = lambda tup: tup[0])[-1][2] + 15
            items_y_start = sorted(bounding_boxes, key = lambda tup: tup[1])[0][1] - 15
            items_y_end = sorted(bounding_boxes, key = lambda tup: tup[1])[-1][1] + sorted(bounding_boxes, key = lambda tup: tup[1])[-1][-1] + 15    
            
            # use a copy to crop the image
            items = img[items_y_start-5:items_y_end+5, items_x_start:items_x_end]
                                  
            # with the location of the items table determined, all other needed parts of the image can be determined relative to it as so:
            top = img[250:items_y_start-125, items_x_start+225:items_x_end]
            
            # get category manager, vendor number, bill to vendor number and vendor name
            bounding_boxes = get_bounding_boxes(top)
            bounding_boxes = [bounding_box for bounding_box in bounding_boxes if (((bounding_box[-2] >= 150) and (bounding_box[-2] <= 510)) and ((bounding_box[-1] >= 35) and (bounding_box[-1] <= 57)))]
            bounding_boxes = sorted(bounding_boxes)
            boxes_dict = neighboring_cells(bounding_boxes, "x")
            left_right = [boxes_dict[key] for key in boxes_dict.keys() if len(boxes_dict[key]) > 1]
            try:
                left = left_right[0]
                right = left_right[1]
                
                crop_left_dim = (sorted(left, key = lambda tup: tup[0])[0][0], sorted(left, key = lambda tup: tup[1])[0][1], sorted(left, key = lambda tup: tup[2])[-1][-2]
                                 ,sorted(left, key = lambda tup: tup[3])[-1][1] + sorted(left, key = lambda tup: tup[3])[-1][-1]
                                 )
                
                crop_right_dim = (sorted(right, key = lambda tup: tup[0])[0][0], sorted(right, key = lambda tup: tup[1])[0][1], sorted(right, key = lambda tup: tup[2])[-1][-2]
                                  , sorted(right, key = lambda tup: tup[3])[-1][1] + sorted(right, key = lambda tup: tup[3])[-1][-1]
                                  )
                
                crop_left = top[crop_left_dim[1]:crop_left_dim[1]+crop_left_dim[-1], crop_left_dim[0]:crop_left_dim[0]+crop_left_dim[-2]]
                crop_right = top[crop_right_dim[1]:crop_right_dim[1]+crop_right_dim[-1], crop_right_dim[0]:crop_right_dim[0]+crop_right_dim[-2]]
                
                left_txt = pytesseract.image_to_string(PIL.Image.fromarray(crop_left), lang = 'eng', config = '--psm 6')
                left_txt = left_txt.split("\n")
                left_txt = list(filter(None, left_txt))
                category_mgr, vendor_num = left_txt[0].replace("_", "").strip(" ").title(), left_txt[2].replace("_", "").strip(" ")
                
                right_txt = pytesseract.image_to_string(PIL.Image.fromarray(crop_right), lang = 'eng', config = '--psm 6')
                right_txt = right_txt.split("\n")
                right_txt = list(filter(None, right_txt))
                bill_to_vendor, vendor_name = right_txt[0].replace("_", "").strip(" "), right_txt[1].replace("_", "").strip(" ")
            except:
                crop_left, crop_right = img[0:1, 0:1], img[0:1, 0:1]
                category_mgr, vendor_num, bill_to_vendor, vendor_name = "", "", "", ""
            
            # extract dates using the previously save date file. use a 'try' statement with an 'except' clause using 'pass' to skip errors
            try:
                dates = cv2.imread(os.path.join(output_path, output_name + "_date.png"))
                img_gray = cv2.cvtColor(dates, cv2.COLOR_BGR2GRAY)
                img_gray = cv2.blur(img_gray, (15, 15))
                ret, thresh = cv2.threshold(img_gray, math.floor(np.average(img_gray)), 255, cv2.THRESH_BINARY_INV)
                dilated = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (10,10)))
                
                contours, _ = cv2.findContours(dilated, cv2.RETR_LIST,cv2.CHAIN_APPROX_SIMPLE)
                new_contours = []
                for contour in contours:
                    if ((cv2.contourArea(contour) < 4000000) and cv2.contourArea(contour) > 2):
                        new_contours.append(contour)
                
                bounding_boxes = [cv2.boundingRect(contour) for contour in new_contours]
                # this next line will return the bounding_boxes for the portion of the image that has dates
                bounding_boxes = [bounding_box for bounding_box in bounding_boxes if (((bounding_box[-2] >= 435) and (bounding_box[-2] <= 520)) and ((bounding_box[-1] >= 180) and (bounding_box[-1] <= 225)))]
                x, y, w, h = bounding_boxes[0][0], bounding_boxes[0][1], bounding_boxes[0][-2], bounding_boxes[0][-1]
                date_box = dates[y:y+h, x:x+w]
                           
                # extract text
                dates = pytesseract.image_to_string(PIL.Image.fromarray(dates), lang = 'eng', config = '--psm 4')
                dates = re.findall("([0-9]+\/+[0-9]+\/[0-9]+)", dates)
                
                if (len(dates) >= 2): # making it >= takes care of cases if somehow more than 2 strings are found using the regex pattern. two variables are assigned either way
                    date_start, date_end = dates[0], dates[1]
                elif len(dates) == 1:
                    date_start, date_end = dates[0], ""
                else:
                    date_start, date_end = "", ""
                    
                del img_gray, ret, thresh, dilated    
                
            except:
                date_box = dates # since date_box is being returned, have it set to dates. Since it is executing as part of the except clause, it won't matter what date_box is anyway so long as some variable is returned
                date_start, date_end = "", ""
                pass
        
            # save the cropped items table image -- useful to for testing
            #cv2.imwrite(os.path.join(output_path, output_name + "_items.png" ), items)
            #date = img[0:, items_x_end:].copy()
            #cv2.imwrite(os.path.join(output_path, output_name + "_date.png"), date)
            #cv2.imwrite(os.path.join(output_path, output_name + "_top.png"), top)
        
            del top 
            return category_mgr, vendor_num, bill_to_vendor, vendor_name, date_start, date_end, items, crop_left, crop_right, date_box
            
        elif template_type == 2:
            
            # want only the bounding boxes which correspond to the cells in the items table that may actually hold information -- height is between 13 and 20 or 35 and 50 and the width is at most 1100
            bounding_boxes = [bounding_box for bounding_box in bounding_boxes if (((bounding_box[-1] >= 23 and bounding_box[-1] <= 35) or (bounding_box[-1] >= 65 and bounding_box[-1] <= 80)) and bounding_box[-2] <= 600)]
            # sort 'bounding_boxes' by y component
            bounding_boxes = sorted(bounding_boxes, key = lambda tup: tup[1])
            
            boxes_dict = neighboring_cells(bounding_boxes, "y")

            # filter for groups of contours that have more than 8 contours in succession
            bounding_boxes = [boxes_dict[key] for key in boxes_dict.keys() if len(boxes_dict[key]) >= 8]
            bounding_boxes = [tup for lst in bounding_boxes for tup in lst]
            
            bounding_boxes = sorted(bounding_boxes)
            boxes_dict = neighboring_cells(bounding_boxes, "x")
            bounding_boxes = [boxes_dict[key] for key in boxes_dict.keys() if len(boxes_dict[key]) >= 5]
            bounding_boxes = [tup for lst in bounding_boxes for tup in lst]
            
            items_x_start = sorted(bounding_boxes, key = lambda tup: tup[0])[0][0] - 15
            items_x_end = sorted(bounding_boxes, key = lambda tup: tup[0])[-1][0] + sorted(bounding_boxes, key = lambda tup: tup[0])[-1][2] + 15
            items_y_start = sorted(bounding_boxes, key = lambda tup: tup[1])[0][1] - 10
            items_y_end = sorted(bounding_boxes, key = lambda tup: tup[1])[-1][1] + sorted(bounding_boxes, key = lambda tup: tup[1])[-1][-1] + 10
            
            items = img[items_y_start:items_y_end, items_x_start:items_x_end]
            top = img[400:items_y_start, items_x_start+250:items_x_end-350]      
            
            bounding_boxes = get_bounding_boxes(top)
            bounding_boxes = [bounding_box for bounding_box in bounding_boxes if (((bounding_box[-2] >= 450) and (bounding_box[-2] <= 710)) and ((bounding_box[-1] >= 30) and (bounding_box[-1] <= 140)))]
            
            bounding_boxes = sorted(bounding_boxes)
            boxes_dict = neighboring_cells(bounding_boxes, "x")
            left_right = [boxes_dict[key] for key in boxes_dict.keys() if len(boxes_dict[key]) >= 2]
            left = left_right[0]
            right = left_right[1]
            
            crop_left_dim = (sorted(left, key = lambda tup: tup[0])[0][0], sorted(left, key = lambda tup: tup[1])[0][1], sorted(left, key = lambda tup: tup[2])[-1][-2]
                             ,sorted(left, key = lambda tup: tup[3])[-1][1] + sorted(left, key = lambda tup: tup[3])[-1][-1]
                             )
            
            crop_right_dim = (sorted(right, key = lambda tup: tup[0])[0][0], sorted(right, key = lambda tup: tup[1])[0][1], sorted(right, key = lambda tup: tup[2])[-1][-2]
                              , sorted(right, key = lambda tup: tup[3])[-1][1] + sorted(right, key = lambda tup: tup[3])[-1][-1]
                              )
            
            crop_left = top[crop_left_dim[1]:crop_left_dim[1]+crop_left_dim[-1], crop_left_dim[0]:crop_left_dim[0]+crop_left_dim[-2]]
            crop_right = top[crop_right_dim[1]:crop_right_dim[1]+crop_right_dim[-1], crop_right_dim[0]:crop_right_dim[0]+crop_right_dim[-2]]
                       
            left_txt = pytesseract.image_to_string(PIL.Image.fromarray(crop_left), lang = 'eng', config = '--psm 6')
            left_txt = left_txt.split("\n")
            left_txt = list(filter(None, left_txt))
            category_mgr, vendor_num = left_txt[0].replace("_", "").replace("—", "").strip(" ").title(), left_txt[2].replace("_", "").replace("—", "").strip(" ")
            
            right_txt = pytesseract.image_to_string(PIL.Image.fromarray(crop_right), lang = 'eng', config = '--psm 6')
            right_txt = right_txt.split("\n")
            right_txt = list(filter(None, right_txt))
            bill_to_vendor, vendor_name = right_txt[0].replace("_", "").replace("—", "").strip(" "), right_txt[1].replace("_", "").replace("—", "").strip(" ")
            
            #cv2.imwrite(os.path.join(output_path, output_name + "_items.png"), items)
            #cv2.imwrite(os.path.join(output_path, output_name + "_top.png"), top)
            
            del top
            return category_mgr, vendor_num, bill_to_vendor, vendor_name, items, crop_left, crop_right
        
        
    #def save_image(self, output_path, file_name, img):
        
        #cv2.imwrite(os.path.join(output_path, file_name), img)
       

class extractItemData():
    
    def __init__(self, img, template_type):
        
        self.img = img
        self.template_type = template_type

    
    def prep_bounding(self):
        
        """Initial prepping for image processing including noise reduction, conversion to black and white, threshholding, etc."""
        
        img = self.img
        img = cv2.fastNlMeansDenoisingColored(img, None, 10, 10, 7, 21)
        img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        _, img = cv2.threshold(img, 240, 255, cv2.THRESH_TOZERO)
        thresh, img_bin = cv2.threshold(img, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        img_bin = 255 - img_bin
    
        #create 3 kernels:  (1) countcol(width) of kernel as 100th of total width 
        #                   (2) vertical kernel to detect all vertical lines of image 
        #                   (3) a 2x2 kernel for morphological purposes
        
        kernel_len = np.array(img).shape[1]//100
        vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, kernel_len))
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        
        del thresh
        return img, img_bin, vertical_kernel, kernel
    

    def vertical_lines(self, img, vertical_kernel, kernel, erode_iter1, erode_iter2, erode_len, erode_width, dilate_iter, dilate_len, dilate_width):
        
        """Use vertical kernel to detect and save the vertical lines in a jpg"""
        
        eroded = cv2.erode(img, vertical_kernel, iterations = erode_iter1)
        dilated = cv2.dilate(eroded, np.ones((dilate_len, dilate_width),np.uint8), iterations = dilate_iter)
        vertical_lines = cv2.erode(dilated, np.ones((erode_len, erode_width),np.uint8), iterations = erode_iter2)
        
        #plotting = plt.imshow(vertical_lines,cmap='gray')
        #plt.show()
    
        del eroded, dilated
        return vertical_lines


    def horizontal_lines(self, img, erode_iter1, erode_iter2, erode_len1, erode_width1, erode_len2, erode_width2, dilate_iter, dilate_len, dilate_width):
        
        """Use horizontal kernel to detect and save the horizontal lines in a jpg"""
        
        eroded = cv2.erode(img, np.ones((erode_width1, erode_len1),np.uint8), iterations = erode_iter1)
        dilated = cv2.dilate(eroded, np.ones((dilate_width, dilate_len),np.uint8), iterations = dilate_iter)
        horizontal_lines = cv2.erode(dilated, np.ones((erode_width2, erode_len2),np.uint8), iterations = erode_iter2)
        
        #plotting = plt.imshow(horizontal_lines,cmap='gray')
        #plt.show()
        
        del eroded, dilated
        return horizontal_lines
    
    
    def sort_contours(self, contours, method = "left-to-right"):
    
        # initialize the reverse flag and sort index
        reverse = False
        i = 0
        
        # handle if we need to sort in reverse
        if method == "right-to-left" or method == "bottom-to-top":
            reverse = True
            
        # handle if we are sorting against the y-coordinate rather than the x-coordinate of the bounding box
        
        if method == "top-to-bottom" or method == "bottom-to-top":
            i = 1
            
        # construct the list of bounding boxes and sort them from top to bottom
        boundingBoxes = [cv2.boundingRect(contour) for contour in contours]
        (contours, boundingBoxes) = zip(*sorted(zip(contours, boundingBoxes), key = lambda b:b[1][i], reverse = reverse))
        
        return (contours, boundingBoxes)


    def bounding_boxes(self, img, kernel, vertical_lines, horizontal_lines, text_removed):
        
        """Get bounding boxes for all cells in the items table. Has several more steps in addition to the 'crop_image' function"""
        
        # used for plotting purposes later on towards the end of this function--useful for testing
        #img_og = cv2.cvtColor(img, cv2.cv2.COLOR_GRAY2RGB)
        
        # Combine horizontal and vertical lines in a new third image, with both having same weight.
        img_vh = cv2.addWeighted(vertical_lines, 0.5, horizontal_lines, 0.5, 0.0)
        rows, cols = img_vh.shape
    
        # shift image -- needed because new horizontal and vertical lines do not overlap with those on the original image
        M = np.float32([[1,0,-5],[0,1,-5]])
        img_vh = cv2.warpAffine(img_vh ,M,(cols,rows))
    
        # Eroding and thesholding the image
        img_vh = cv2.erode(~img_vh, kernel, iterations = 2)
        thresh, img_vh = cv2.threshold(img_vh, 128, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        
        bitxor = cv2.bitwise_xor(img, img_vh)
        bitnot = cv2.bitwise_not(bitxor)
        
        # plotting = plt.imshow(bitnot,cmap='gray')
        # plt.show()
        
        # Detect contours for following box detection
        contours, hierarchy = cv2.findContours(img_vh, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
        
        # Sort all the contours by top to bottom.
        contours, boundingBoxes = extractItemData(self.img, self.template_type).sort_contours(contours, method = "top-to-bottom")
                
        # Creating a list of heights for all detected boxes
        heights = [boundingBoxes[i][3] for i in range(len(boundingBoxes))]
        
        # Get mean of heights
        height_mean = np.mean(heights)
        
        # Create list box to store all boxes in  
        box_list = []
        
        # Get position (x,y), width and height for every contour and show the contour on image
        for contour in contours:
            x, y, w, h = cv2.boundingRect(contour)
            
            if self.template_type == 1:
            # add contours to 'box_list' if the following conditions are met for width and height (this will restrict to cells in the table that may contain data)
                if ((w > 100) and (w < 700) and (h > 40) and (h < 900)):
                    
                    #overlay the identified bounding boxes on the original image -- useful for testing
                    #print(str((x,y,w,h)))
                    #img_plot = cv2.rectangle(img_og, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    #plt.imshow(img_plot, cmap = 'gray')
                    #plt.show()
                
                    box_list.append([x, y, w, h])
                
                # this condition is to white out text held inside potential cells determined by contours to later dilate horizontal or vertical lines if necessary
                # the indicated width and height in the 'elif' condition are so that the entire image containing the table with data is excluded from being whited out
                elif not ((w > 700) and (h > 600)):
                    text_removed[y:y+h-1,x:x+w-1] = [255,255,255]
                    
            elif self.template_type == 2:
                # condition with the x is so that the cell containing 'scan dates' is not used.
                if ((not ((x >= 1370 and x <= 1390) and (w >= 345 and w <= 365))) and ((w > 100 and w < 600) and (h > 10 and h < 85))):
                    #print(str((x,y,w,h)))
                    
                    # overlay the identified bounding boxes on the original image -- useful for testing
                    # img_plot = cv2.rectangle(img_og, (x, y), (x + w, y + h), (255, 0, 0), 2)
                    # plt.imshow(img_plot, cmap = 'gray')
                    # plt.show()
                    
                    box_list.append([x, y, w, h])
                    
                    #time.sleep(1)
                elif not ((w > 1700) and (h > 975)):
                    text_removed[y:y+h-1,x:x+w-1] = [255, 255, 255]
        
        del img_vh, thresh, bitxor
        gc.collect()
        
        return bitnot, box_list, text_removed, height_mean
        
    
    def final_bounding_boxes(self, box_list, height_mean):
        
        # Creating two lists to define row and column in which cell is located
        row, column, final_boxes, j, column_count = [], [], [], 0, 0
        
        # Sorting the boxes to their respective row and column
        for i in range(len(box_list)):    
                
            if i == 0:
                column.append(box_list[i])
                previous = box_list[i]    
            else:
                if (box_list[i][1] <= previous[1] + height_mean/2):
                    column.append(box_list[i])
                    previous = box_list[i]            
                    
                    if (i == len(box_list) - 1):
                        row.append(column)        
                else:
                    row.append(column) 
                    column=[]
                    previous = box_list[i]
                    column.append(box_list[i])
                    
        # calculating maximum number of cells
        for i in range(len(row)):
            column_count = len(row[i])
            if column_count > column_count:
                column_count = column_count

        # Retrieving the center of each column
        center = [int(row[i][j][0] + row[i][j][2]/2) for j in range(len(row[i])) if row[0]]
        center = np.array(center)
        center.sort()

        # arrange boxes in respective order to distance to column's center
        for i in range(len(row)):
            lis = []
            
            for k in range(column_count):
                lis.append([])
                
            for j in range(len(row[i])):
                diff = abs(center - (row[i][j][0] + row[i][j][2]/4))
                minimum = min(diff)
                indexing = list(diff).index(minimum)
                lis[indexing].append(row[i][j])
                
            final_boxes.append(lis)
        
        return row, column_count, final_boxes
        
    
    def extract_item_data(self, bitnot_img, row, column_count, final_boxes):
        
        """Perform OCR for text extraction using the bit not image and final set of bounding boxes"""
        
        outer = []    
    
        # from every single image-based cell/box the strings are extracted via pytesseract and stored in a list
        if self.template_type == 1:
            start = 1 # rows (not inclunding headers) for template 1 start on 1
        elif self.template_type == 2:
            start = 2 # rows (not including headers) for template 2 start on 2
        
        for i in range(start, len(final_boxes)):
            
            for j in range(len(final_boxes[i])):
                 
                if (len(final_boxes[i][j]) == 0):
                    outer.append(' ')
                
                else:
                    
                    for k in range(len(final_boxes[i][j])):
                        
                        y, x, w, h = final_boxes[i][j][k][0], final_boxes[i][j][k][1], final_boxes[i][j][k][2], final_boxes[i][j][k][3]
                        
                        if self.template_type == 1:
                            
                            if j == 4:
                                y += 40 #crop some more to get rid of dolar sign
                                w -= 30 #crop some more to get rid of vertical border to the right of amount
                                
                            final_img = bitnot_img[x:x+h, y:y+w]
                            
                            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 1))
                            resizing = cv2.resize(final_img, None, fx = 2, fy = 2, interpolation = cv2.INTER_CUBIC)
                            dilation = cv2.dilate(resizing, kernel, iterations = 1)
                            erosion = cv2.erode(dilation, kernel, iterations = 2)
                            
                            #plot the cell found in 'final_img' after transforming -- useful for checking what Python sees before applying OCR
                            #plt.imshow(erosion, cmap = 'gray')
                            #plt.show()
        
                            out = pytesseract.image_to_string(erosion, lang = 'eng', config = '--psm 12')
                            out = out.replace("$", "").replace("_", "").replace("|", "").replace("!", "").strip()
                            
                            if (j == 4):
                                re_bool = re.search('(\d*\.*\d+)', out)
                                if re_bool:
                                    out = re_bool.group(1)
                        
                        elif self.template_type == 2:
                            
                            if j == 0:
                                w += 1
                            elif j == 5:
                                y += 30
                                w -= 30
                            
                            final_img = bitnot_img[x:x+h, y:y+w]
                            
                            kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 1))
                            resizing = cv2.resize(final_img, None, fx = 2, fy = 2, interpolation = cv2.INTER_CUBIC)
                            dilation = cv2.dilate(resizing, kernel, iterations = 1)
                            erosion = cv2.erode(dilation, kernel, iterations = 2)
                            
                            #plt.imshow(erosion, cmap = 'gray')
                            #plt.show()
                            
                            out = pytesseract.image_to_string(erosion, lang = 'eng', config = '--psm 12')
                            out = out.replace("$", "").replace("_", "").replace("|", "").replace("!", "").replace("\n", "—").strip()
                            
                            if (j == 5):
                                re_bool = re.search('(\d*\.*\d+)', out)
                                if re_bool:
                                    out = re_bool.group(1)
                        
                        #print(i,j,k,out)
                        if (j == 0):
                            out = out.replace("\n", "").replace("—", "").strip()
    
                    outer.append(out)
                    
        #Creating a dataframe of the generated OCR list
        arr = np.array(outer)
        frame = pd.DataFrame(arr.reshape(len(row) - start, column_count))
        
        #remove empty rows from frame to show only those that have data
        frame = frame[(frame[0].str.strip().astype(bool)) & (frame[2].str.strip().astype(bool))]
    
        del final_img, kernel, resizing, dilation, erosion, out
        gc.collect()
        
        return frame



def time_convert(start, end):
    
    """Function to track how long the entire program took to execute"""
    
    duration = round(end - start, 2)
    
    if duration < 60:
        print('\n** Duration: ' + str(duration) + ' seconds.')
    else:
        duration = duration/60
        min_part = int(duration//1)
        sec_part = int(round((duration % 1) * 60, 0))
        if min_part == 1:
            print('\n ** Duration: ' + str(min_part) + ' minute, ' + str(sec_part) + ' seconds')
        else:
            print('\n ** Duration: ' + str(min_part) + ' minutes, ' + str(sec_part) + ' seconds')



def create_directory(path, name = ""):
    
    """This function will be used to create a folder than will contain all associated images of the image with text to be extracted"""
    
    # 'name' in blank in case a path only needs to be created without a folder
    directory = os.path.join(path, name)
    directory = directory.replace("\\", "/")

    if not os.path.exists(directory):
        os.mkdir(directory)
    else:
        incr = 1
        while os.path.exists(directory):
            
            name_pattern = re.compile("^.+\(+\d+\)$")
            
            if bool(re.match(name_pattern, directory)):
                # if the file name ends with a number in parenthesis, extract the number and add 1 to it
                incr = directory.split(re.split("\(+\d+\)", directory)[0])[-1]
                directory = re.split("\(+\d+\)", directory)[0]
                incr = incr.replace("(", "").replace(")", "")
                incr = str(int(incr) + 1)
                directory = directory + "(" + incr + ")"

            else:
                if not (name == ""):
                    name = name + " (" + str(incr) + ")"
                    directory = directory.rsplit("/", 1)[0]
                else:
                    name = directory.rsplit("/")[-1]
                    directory = directory.rsplit(name)[0]
                    name = name.replace("\\", "").strip(" ") + " (" + str(incr) + ")"
                    # print(name, directory)
                
                directory = os.path.join(directory, name)

        os.mkdir(directory)
    
    return directory
    

def create_workbook(data_frame_list, path, name):
    
    """Take dataframe output and write data to Excel workbook"""
    
    wb_path = os.path.join(path, name + '.xlsx')
    writer = pd.ExcelWriter(wb_path, engine = 'xlsxwriter')
    
    row_start = 0
    incr = 1
    for frame_index, data_frame in enumerate(data_frame_list[0]):
        
        if data_frame.shape[1] == 10:
            if frame_index == 0:
                data_frame.to_excel(writer, sheet_name = 'ExtractedText', startrow = 1, index = False, header = False)
                row_start += 1
            else:
                row_start += data_frame_list[0][frame_index - 1].shape[0]
                data_frame.to_excel(writer, sheet_name = 'ExtractedText', startrow = row_start, index = False, header = False)
        else:
            data_frame.to_excel(writer, sheet_name = 'Sheet' + str(incr), index = False, header = False)
            
    
    # write headers
    workbook  = writer.book
    worksheet = writer.sheets['ExtractedText']
    col_num = 0
    headers = ["PageNumber", "ItemNumber", "ItemDescription", "ScanDiscount", "CategoryManager", "VendorNumber", "BillToVendor", "VendorName", "DateStart", "DateEnd", "AdGroup", "SuggestedFileName"]
    headers_format = workbook.add_format({'bold': True})
    
    for header in headers:
        worksheet.write(0, col_num, header, headers_format)
        col_num += 1
    
    writer.save()
    
    return wb_path


def auto_fit(file_path):
    
    """Simulates autofit in Excel"""
    
    wb = op.load_workbook(filename = file_path)
    
    for worksheet in wb:
        for col in worksheet.columns:
            max_length = 0 
            column = col[0].column_letter
            for cell in col:
                if cell.coordinate in worksheet.merged_cells:
                    continue
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column].width = adjusted_width
       
    sheet = wb['ExtractedText']
    row_count = sheet.max_row
    
    # add 'All' for 'AdGroup' column and alignment column A left
    for row in range(0, row_count):
        sheet.cell(column = 11, row = row + 2, value = "All")
        sheet.cell(column = 1, row = row + 1).alignment = op.styles.Alignment(horizontal = 'left')
        
    wb.save(file_path)

def extract(output_fmt, output_path, pdf_path, img, page_num):
    
    # 'output_name' will be the name of the file given by 'pdf_path' retrieved using string manipulation to get the file name
    output_name = pdf_path.rsplit("/", 1)[-1].split(".")[0] + "_page" + str(page_num + 1)
    output_name_complete = output_name + "." + output_fmt
    
    # create a folder that will store all images (cropped, items, vendor, etc.) for the page currently being worked on
    path_save_folder = create_directory(output_path, output_name)
    path_save = os.path.join(path_save_folder, output_name_complete)
    
    img.rotate(pytesseract.image_to_osd(img, output_type = pytesseract.Output.DICT)['orientation'], expand = True).save(path_save)
    print("\n\tPage " + str(page_num + 1) + " saved in '" + output_fmt + "' format.")
    
    del img
    gc.collect()
    
    prep_instance = imagePrep()
    
    # remove skew and determine the template type
    img_path = prep_instance.remove_skew(img_path = path_save, save_path = path_save_folder, page_num = page_num)
    print("\n\tImage path: " + img_path)
    type_result = prep_instance.template_type(img_path)
    
    # include this while loop in case the wrong template type was determined (it happens occasionally)
    counter = 1
    while counter <= 2:
        try:
            # if the template type is 1, then crop the image as follows, otherwise use the steps for template type 2
            if type_result == 1:
                category_mgr, vendor_num, bill_to_vendor, vendor_name, date_start, date_end, items, crop_left, crop_right, date_box = prep_instance.crop_image(output_name, path_save_folder, img_path, type_result)
                
            elif type_result == 2:
                category_mgr, vendor_num, bill_to_vendor, vendor_name, items, crop_left, crop_right = prep_instance.crop_image(output_name, path_save_folder, img_path, type_result)
                
        except:
            if ((type_result == 1) and (counter < 2)):
                print("\n\tTemplate type changed from 1 to 2")
                type_result = 2
                counter += 1
            elif ((type_result == 2) and (counter < 2)):
                print("\n\tTemplate type changed from 2 to 1")
                type_result = 1
                counter += 1
            elif counter == 2: # once the counter reaches 3, it means the image is neither a type1 or type2 template, so set it equal to type3 and break out of the loop
                print("\n\tTemplate type set to 3")    
                type_result = 3
                break
            continue
        else:
            break
    
    print("\n\tImage cropped.")
    
    # proceed with the following steps if the type_result is 1 or 2, type_result 3 will follow the 'else' statement
    try:
        if ((type_result == 1) or (type_result == 2)):
              
            # create 'removed_text' which initially is just a copy of the items after cropping
            # this variable will be used later if the original 'items' image is determined to have fewer than the appropriate number of cells
            removed_text = items
            
            # create an extract instance, followed by kernels, vertical/horizontal line kernels, and the image of items to be used for extraction
            extract_instance = extractItemData(items, type_result)
            image, image_bin, vKern, kern = extract_instance.prep_bounding()
            vLines = extract_instance.vertical_lines(image_bin, vKern, kern, erode_iter1 = 4, erode_iter2 = 2, erode_len = 25, erode_width = 2, dilate_iter = 4, dilate_len = 25, dilate_width = 2)
            hLines = extract_instance.horizontal_lines(image_bin, erode_iter1 = 10, erode_iter2 = 2, erode_len1 = 5, erode_width1 = 1, erode_len2 = 39, erode_width2 = 2, dilate_iter = 7, dilate_len = 40, dilate_width = 2)
            img_bitnot, list_box, removed_text, hgt_mean = extract_instance.bounding_boxes(image, kern, vLines, hLines, removed_text)
        
            if (((type_result == 1) and (len(list_box) < 70)) or ((type_result == 2) and (len(list_box) < 40))): 
                # for type 1: if all contours in a type 1 template are properly detected, it should have exactly 70 cells (5 colums by 14 rows). otherwise, the image needs adjusting--'removed_text' is used
                # for type 2: this is an arbitrary choice. type 2 templates always have 8 columns, but vary in row count. this assumes there are only 5 rows detected -- should be a good lower limit
                print("\n\tMissing boxes detected. Adjusting...")
                image, image_bin, vKern, kern = extractItemData(removed_text, type_result).prep_bounding()
                vLines = extract_instance.vertical_lines(image_bin, vKern, kern, erode_iter1 = 4, erode_iter2 = 1, erode_len = 60, erode_width = 2, dilate_iter = 7, dilate_len = 60, dilate_width = 2)
                hLines = extract_instance.horizontal_lines(image_bin, erode_iter1 = 10, erode_iter2 = 8, erode_len1 = 5, erode_width1 = 1, erode_len2 = 41, erode_width2 = 2, dilate_iter = 12, dilate_len = 45, dilate_width = 2)
                img_bitnot, list_box, removed_text, hgt_mean = extract_instance.bounding_boxes(image, kern, vLines, hLines, removed_text)
           
            # get the final bounding boxes and extract text from 'img_bitnot'
            rw, col_count, boxes_final = extract_instance.final_bounding_boxes(list_box, hgt_mean)
            extracted_data = extract_instance.extract_item_data(img_bitnot, rw, col_count, boxes_final)
        
            # add columns to the 'extracted_data' dataframe with the following details
            extracted_data["PageNum"] = [page_num + 1] * extracted_data.shape[0]
            extracted_data["CategoryManager"] = [category_mgr] * extracted_data.shape[0]
            extracted_data["VendorNum"] = [vendor_num] * extracted_data.shape[0]
            extracted_data["BillToVendor"] = [bill_to_vendor] * extracted_data.shape[0]
            extracted_data["VendorName"] = [vendor_name] * extracted_data.shape[0]
        
            if type_result == 1:
                # add date columns to dataframe if template type is 1
                extracted_data["DateStart"] = [date_start] * extracted_data.shape[0]
                extracted_data["DateEnd"] = [date_end] * extracted_data.shape[0]
                
                # rename columns if there are 12 of them (needs to be 12 otherwise it means pytesseract couldn't detect the correct number of columns)
                if extracted_data.shape[1] == 12:
                    extracted_data.rename(columns = {0: 'S&FCode', 1: 'UPCCode', 2: 'ItemDescription', 3: 'Size', 4: 'ScanDiscount'}, inplace = True)
                    extracted_data = extracted_data[["PageNum", "S&FCode", "ItemDescription", "ScanDiscount", "CategoryManager", "VendorNum", "BillToVendor", "VendorName", "DateStart", "DateEnd"]]
                    print("\t\nColumns renamed")
            
            elif type_result == 2:
                if extracted_data.shape[1] == 13:
                    # rename columns if there are 13 of them (needs to be 13 otherwise it means pytesseract couldn't detect the correct number of columns)
                    extracted_data.rename(columns = {0: 'S&FCode', 1: 'UPCCode', 2: 'ItemDescription', 3: 'CasePack', 4: 'Size', 5: 'ScanDiscount', 6: 'DateStart', 7: 'DateEnd'}, inplace = True)
                    extracted_data = extracted_data[["PageNum", "S&FCode", "ItemDescription", "ScanDiscount", "CategoryManager", "VendorNum", "BillToVendor", "VendorName", "DateStart", "DateEnd"]]
                    print("\t\nColumns renamed")
            
            if type_result == 1:
                del date_start, date_end, date_box
            
            del removed_text, image, image_bin, vKern, kern, vLines, hLines, img_bitnot, category_mgr, vendor_num, bill_to_vendor, vendor_name, items, crop_left, crop_right
            
            gc.collect()
    
        else:
            img = cv2.imread(img_path)
            txt = pytesseract.image_to_string(img, lang = 'eng', config = '--psm 6')
            # split the txt to get extracted text equivalent to line by line format
            txt = txt.split("\n")
            txt = [re.search(r"(\d+)(.*)", text).group(0) for text in txt if bool(re.match(r"(\d+)(.*)", text))]
            txt = [list(filter(None, re.split(r"(^[\d]+)", text))) for text in txt]
            extracted_data = pd.DataFrame(txt)
        
            del img, txt
    except:
        data = [page_num, "", "COULD NOT EXTRACT TEXT", "", "", "", "", "", "", ""]
        cols = ["PageNum", "S&FCode", "ItemDescription", "ScanDiscount", "CategoryManager", "VendorNum", "BillToVendor", "VendorName", "DateStart", "DateEnd"]
        extracted_data =pd.DataFrame(data, columns = cols)
        
    gc.collect()
    
    return extracted_data


if __name__ == "__main__":

    pdf_path = r"C:/Users/erik7/Documents/Late Scans for Testing/scans_ass_load.pdf"
    output_fmt = 'jpeg'
    img_dpi = 300 #dpi should be at least 300 for accurate extraction of text -- higher dpi impacts performance negatively
    pop_path = r"C:\Users\erik7\Downloads\poppler-0.90.1\bin"
    output_path = r"C:\Users\erik7\Downloads"
       
    start = time.time()
    
    converted_path = create_directory(output_path, "converted_images")
    print("\nConverting PDF to " + str(output_fmt) + "...")
    converted = imagePrep().convert_from_pdf(pdf_path, output_fmt, img_dpi, pop_path, converted_path)
    print("\nPDF converted to " + str(output_fmt))

    results = [] 
    iterable = [[output_fmt, output_path, pdf_path, img, page_num] for page_num, img in enumerate(converted[:2])]
    p = multiprocessing.Pool()
    r = p.starmap(extract, iterable)
    results.append(r)
    p.close()
    #del img, data_extracted
    #gc.collect()
    
    # for page_num, img in enumerate(converted[:10]):
    #     print("\nPage " + str(page_num + 1) + " of " + str(len(converted)))
        
    #     data_extracted = extract(output_fmt, output_path, pdf_path, img, page_num)
    #     results.append(data_extracted)
        
    #     del img, data_extracted
    #     gc.collect()
               
    print("\n**PROCESS COMPLETED SUCCESSFULLY")
    
    end = time.time()
    time_convert(start, end)  
    
    #workbook_path = create_workbook(results, r"C:\Users\erik7\Downloads", "results")
    #auto_fit(workbook_path)